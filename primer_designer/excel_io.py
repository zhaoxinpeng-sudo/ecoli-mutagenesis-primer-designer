from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


INPUT_COLUMNS = ["mutation"]
OUTPUT_COLUMNS = [
    "mutation", "primer_name", "direction", "primer_sequence_5to3",
    "length", "tm", "gc", "overlap_length", "failure_reason",
]
OUTPUT_COLUMN_LABELS = {
    "mutation": "突变",
    "primer_name": "引物名称",
    "direction": "方向",
    "primer_sequence_5to3": "引物序列（5′→3′）",
    "length": "长度（nt）",
    "tm": "Tm（°C）",
    "gc": "GC（%）",
    "overlap_length": "共享区长度（bp）",
    "failure_reason": "失败原因",
}


def read_mutations(file) -> pd.DataFrame:
    frame = pd.read_excel(file, dtype=str)
    missing = [column for column in INPUT_COLUMNS if column not in frame.columns]
    if missing:
        raise ValueError(f"Excel 缺少必需列：{', '.join(missing)}")
    frame = frame[INPUT_COLUMNS].fillna("")
    frame["mutation"] = frame["mutation"].str.strip()
    frame = frame[frame["mutation"] != ""].reset_index(drop=True)
    if frame.empty:
        raise ValueError("Excel 中没有可用的突变记录")
    return frame


def _write_styled_excel(frame: pd.DataFrame, sheet_name: str, merged_ranges: list[str] | None = None) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        frame.to_excel(writer, index=False, sheet_name=sheet_name)
        sheet = writer.book[sheet_name]
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        fill = PatternFill("solid", fgColor="0F766E")
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        for column in sheet.columns:
            letter = column[0].column_letter
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 45)
            sheet.column_dimensions[letter].width = max(width, 12)
        for cell_range in merged_ranges or []:
            sheet.merge_cells(cell_range)
            sheet[cell_range.split(":")[0]].alignment = Alignment(horizontal="center", vertical="center")
    return output.getvalue()


def template_bytes() -> bytes:
    return _write_styled_excel(pd.DataFrame([["A123V"]], columns=INPUT_COLUMNS), "突变输入")


def flatten_results(records: list[dict]) -> list[dict]:
    """Convert one design record into two order-ready primer rows."""
    rows: list[dict] = []
    for record in records:
        common = {"mutation": record["mutation"]}
        if record["status"] != "成功":
            rows.append({
                **common, "primer_name": "", "direction": "", "primer_sequence_5to3": "",
                "length": None, "tm": None, "gc": None, "overlap_length": None,
                "failure_reason": record["message"],
            })
            continue
        rows.extend([
            {
                **common, "primer_name": record["forward_name"], "direction": "上游引物",
                "primer_sequence_5to3": record["forward_primer"], "length": record["forward_length"],
                "tm": record["forward_tm"], "gc": record["forward_gc"],
                "overlap_length": record["overlap_length"], "failure_reason": "",
            },
            {
                **common, "primer_name": record["reverse_name"], "direction": "下游引物",
                "primer_sequence_5to3": record["reverse_primer"], "length": record["reverse_length"],
                "tm": record["reverse_tm"], "gc": record["reverse_gc"],
                "overlap_length": record["overlap_length"], "failure_reason": "",
            },
        ])
    return rows


def results_bytes(records: list[dict]) -> bytes:
    rows = flatten_results(records)
    merged_ranges: list[str] = []
    excel_row = 2
    for record in records:
        if record["status"] == "成功":
            merged_ranges.append(f"A{excel_row}:A{excel_row + 1}")
            excel_row += 2
        else:
            excel_row += 1
    frame = pd.DataFrame(rows, columns=OUTPUT_COLUMNS).rename(columns=OUTPUT_COLUMN_LABELS)
    return _write_styled_excel(frame, "设计结果", merged_ranges)
